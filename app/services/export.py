import csv
import io
from datetime import datetime
from typing import Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.payment import Payment, PaymentStatus
from app.models.user import User
from app.models.vpn_key import VpnKey
from app.utils.log import log


class ExportService:
    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def export_users(self, fmt: str = "csv") -> bytes:
        """Export all users with subscription/payment counts."""
        result = await self.session.execute(select(User).order_by(User.id))
        users = list(result.scalars().all())

        headers = [
            "Telegram ID", "Full Name", "Username", "Language",
            "Balance", "Banned", "Auto Renew",
            "Created At", "Subscriptions Count", "Payments Count",
        ]
        rows = []
        for u in users:
            rows.append([
                u.id,
                u.full_name or "",
                u.username or "",
                u.language or "",
                float(u.balance or 0),
                "Yes" if u.is_banned else "No",
                "Yes" if u.autorenew else "No",
                u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "",
                len(u.vpn_keys),
                len(u.payments),
            ])
        return self._to_bytes(headers, rows, fmt, "users")

    async def export_payments(
        self,
        fmt: str = "csv",
        status: Optional[str] = None,
        payment_type: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
    ) -> bytes:
        """Export payments with optional filters."""
        stmt = select(Payment).order_by(Payment.created_at.desc())
        if status:
            stmt = stmt.where(Payment.status == status)
        if payment_type:
            stmt = stmt.where(Payment.payment_type == payment_type)
        if date_from:
            try:
                df = datetime.strptime(date_from, "%Y-%m-%d")
                stmt = stmt.where(Payment.created_at >= df)
            except ValueError:
                pass
        if date_to:
            try:
                dt = datetime.strptime(date_to, "%Y-%m-%d")
                stmt = stmt.where(Payment.created_at < dt)
            except ValueError:
                pass

        result = await self.session.execute(stmt)
        payments = list(result.scalars().all())

        headers = [
            "ID", "User ID", "Provider", "Type", "Amount",
            "Currency", "Status", "External ID", "Created At",
        ]
        rows = []
        for p in payments:
            rows.append([
                p.id,
                p.user_id,
                p.provider or "",
                str(p.payment_type) if p.payment_type else "",
                float(p.amount or 0),
                p.currency or "",
                p.status or "",
                p.external_id or "",
                p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "",
            ])
        return self._to_bytes(headers, rows, fmt, "payments")

    async def export_subscriptions(self, fmt: str = "csv") -> bytes:
        """Export VPN keys (subscriptions) with plan info."""
        result = await self.session.execute(
            select(VpnKey).order_by(VpnKey.id.desc())
        )
        keys = list(result.scalars().all())

        headers = [
            "ID", "User ID", "Status", "Plan ID",
            "Expires At", "Access URL", "Created At",
        ]
        rows = []
        for k in keys:
            rows.append([
                k.id,
                k.user_id,
                k.status or "",
                k.plan_id or "",
                k.expires_at.strftime("%Y-%m-%d %H:%M") if k.expires_at else "",
                k.access_url or "",
                k.created_at.strftime("%Y-%m-%d %H:%M") if k.created_at else "",
            ])
        return self._to_bytes(headers, rows, fmt, "subscriptions")

    def _to_bytes(self, headers: list, rows: list, fmt: str, name: str) -> bytes:
        if fmt == "xlsx":
            return self._to_xlsx(headers, rows, name)
        return self._to_csv(headers, rows)

    def _to_csv(self, headers: list, rows: list) -> bytes:
        buf = io.StringIO()
        writer = csv.writer(buf, dialect="excel", lineterminator="\r\n")
        writer.writerow(headers)
        writer.writerows(rows)
        # UTF-8 BOM for Excel compatibility
        return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")

    def _to_xlsx(self, headers: list, rows: list, sheet_name: str) -> bytes:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="00d4aa", end_color="00d4aa", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for r_idx, row in enumerate(rows, 2):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                max_length = 0
                for row in range(1, len(rows) + 2):
                    cell_val = ws.cell(row=row, column=col).value
                    if cell_val:
                        max_length = max(max_length, len(str(cell_val)))
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_length + 2, 50)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.getvalue()
        except ImportError:
            log.warning("openpyxl not installed, falling back to CSV")
            return self._to_csv(headers, rows)
